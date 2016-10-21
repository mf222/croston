from xlrd import open_workbook
import io
import numpy as n
from producto_punto.models import *
import xlsxwriter
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.chart import Series, Reference, BubbleChart
import os
from datetime import datetime, timedelta
import random
from scipy.stats import norm


def WriteToExcel(data, leadtimes):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()  # Sheet con datos
    ws2 = wb.create_sheet(title="Croston")  # Sheet con resultado simulacion
    ws3 = wb.create_sheet(title="p vs NS")  # Sheet para el gráfico

    # Definir aqui el numero de veces que correra
    NUMERO_RUNS = 100
    NUMERO_SIMU = 100

    # Para determinar cuanto es el intervalo de días que se simula
    min_date = None
    max_date = None

    # Diccionario para los datos de la simulación
    data_simulation = dict()

    # Lista para el grafico
    plot_data = []

    actual_row = 7
    for k, v in data.items():
        # Solo se simulan datos que tengan más de 1 pedido
        pedidos = len(v)
        if pedidos > 1:
            data_elements = []
            alldda = []
            alldates = []

            for element in v:
                dda = 0
                if element[2] > 0:
                    dda = element[2]
                alldda.append(dda)
                alldates.append(datetime.strptime(element[1], '%d-%m-%Y'))

            # Para determinar los periodos e intervalos
            alldates.sort()
            difference_days = []
            for i in range(0, len(alldates)-1):
                difference_days.append((alldates[i+1]-alldates[i]).days)
            period = (alldates[-1]-alldates[0]).days
            m_interval = n.mean(difference_days)

            # Para determinar demanda media y desv. estandar
            m_dda = n.mean(alldda)
            std_dda = n.std(alldda)

            if m_interval == 0:
                m_interval = (max_date-min_date).days

            # CROSTON METRICAS
            p_croston = m_interval
            q_croston = 1
            z_croston = m_dda

            y_t = []
            y_t.append(z_croston/p_croston)
            y_t_d = []
            y_t_d_pow = []

            e_mu = []
            e_mu_pow = []

            Mn = []
            Mn.append(std_dda)

            Rt = []
            Rt.append(m_dda+m_dda*std_dda)

            shortage = []
            Mn_temp = []

            alpha = 0.1
            actual_dda = 0
            actual_day = alldates[0]

            for i in range(period):
                e_val = 0
                y_t_d_val = 0 - z_croston/p_croston
                mn_val = Mn[i]
                rt_val = 0
                shortage_val = 0
                mn_temp = 0
                if actual_day == v[actual_dda][1]:
                    e_val = v[actual_dda][2] - z_croston
                    y_t_d_val = v[actual_dda][2] - z_croston/p_croston
                    p_croston = (p_croston*(1-alpha))+(alpha*q_croston)
                    temp = z_croston
                    z_croston = (z_croston*(1-alpha))+(alpha*v[actual_dda][2])
                    q_croston = 0
                    mn_val = Mn[i]*(1-Mn[i]) + Mn[i]*abs(y_t_d_val)
                    actual_dda += 1
                    mn_temp = Mn[i]
                if Rt[i] < v[actual_dda][2]:
                    shortage_val = 1
                q_croston += 1
                actual_day = actual_day + timedelta(days=1)
                y_t.append(z_croston/p_croston)
                e_mu.append(e_val)
                e_mu_pow.append(e_val**2)
                y_t_d.append(y_t_d_val)
                y_t_d_pow.append(y_t_d_val**2)
                Mn.append(mn_val)
                Rt.append(n.floor(z_croston+z_croston*mn_val))
                shortage.append(shortage_val)
                Mn_temp.append(mn_temp)

            e_nzero = n.trim_zeros(e_mu_pow)
            mn_nzero = n.trim_zeros(Mn_temp)

            if len(e_nzero) == 0:
                e_nzero.append(0)
            if len(mn_nzero) == 0:
                mn_nzero.append(0)

            data_elements.append(pedidos)  # TOTAL PEDIDOS
            data_elements.append(m_interval)  # INTERVALO
            data_elements.append(m_dda)  # DEMANDA MEDIA
            data_elements.append(y_t[-1])  # MU MEAN - MU
            data_elements.append(Mn[-1])  # SIGMA
            data_elements.append(p_croston)  # P

            if k in leadtimes:
                data_elements.append(leadtimes[k])
            else:
                data_elements.append(None)

            data_simulation[k] = data_elements
            actual_row += 1

            if min_date is None:
                min_date = alldates[0]
            elif alldates[0] < min_date:
                min_date = alldates[0]

            if max_date is None:
                max_date = alldates[-1]
            elif alldates[-1] > max_date:
                max_date = alldates[-1]

            ws.cell(row=actual_row, column=1, value=k)
            ws.cell(row=actual_row, column=2, value=n.sum(alldda))
            ws.cell(row=actual_row, column=3, value=m_dda)
            ws.cell(row=actual_row, column=4, value=std_dda)
            ws.cell(row=actual_row, column=5, value=m_interval)
            ws.cell(row=actual_row, column=6, value=alldates[0])
            ws.cell(row=actual_row, column=7, value=alldates[-1])
            ws.cell(row=actual_row, column=8, value=period)
            ws.cell(row=actual_row, column=9, value=p_croston)
            ws.cell(row=actual_row, column=10, value=z_croston)
            ws.cell(row=actual_row, column=11, value=n.sum(y_t))
            ws.cell(row=actual_row, column=12, value=n.mean(y_t))
            ws.cell(row=actual_row, column=13, value=max(Rt))
            ws.cell(row=actual_row, column=14, value=n.sum(shortage))
            ws.cell(row=actual_row, column=15, value=n.mean(Rt))
            ws.cell(row=actual_row, column=16, value=n.mean(y_t_d_pow))
            ws.cell(row=actual_row, column=17, value=n.mean(e_nzero))
            ws.cell(row=actual_row, column=18, value=n.mean(mn_nzero))

    total_period = (max_date-min_date).days

    # INICIO SIMULACION
    print("INICIO SIMULACION")
    actual_row = 8
    for k, v in data_simulation.items():
        ALL_NS = []
        ALL_R_0 = []
        for simulations in range(NUMERO_SIMU):
            NS = 0
            R_0 = 0
            while NS < 0.8:
                dda = []
                dda_desp = []
                leadtime = 0
                if v[6] is not None:
                    leadtime = v[6]
                R = []
                R_0 += 1
                R.append(R_0)
                y_more_R = []
                for i in range(NUMERO_RUNS):
                    selector = 1/(v[5])
                    y = 0
                    if selector <= random.random():
                        y = n.floor(norm.ppf(random.random(), v[3], v[4]))
                        if y < 1 or y != y:
                            y = 1
                    dda.append(y)
                for i in range(NUMERO_RUNS):
                    desplazamiento = i - leadtime
                    if desplazamiento < 0:
                        dda_desp.append(0)
                    else:
                        dda_desp.append(dda[i])
                for i in range(1, NUMERO_RUNS):
                    R_i = R[i-1]-dda[i]+dda_desp[i]
                    R.append(R_i)

                for i in range(NUMERO_RUNS):
                    ymr = 0
                    if dda[i] > R[i]:
                        ymr = 1
                    y_more_R.append(ymr)

                numero_dda = 0
                for i in range(len(dda)):
                    if dda[i] != 0:
                        numero_dda += 1
                NS = 1 - (n.sum(y_more_R)/numero_dda)
            ALL_NS.append(NS)
            ALL_R_0.append(R_0)

        ns_mean = n.mean(NS)
        ws2.cell(row=actual_row, column=1, value=k)
        ws2.cell(row=actual_row, column=2, value=v[3])
        ws2.cell(row=actual_row, column=3, value=v[4])
        ws2.cell(row=actual_row, column=4, value=v[5])
        ws2.cell(row=actual_row, column=5, value=n.mean(R_0))
        ws2.cell(row=actual_row, column=6, value=ns_mean)
        plot_data.append((1, ns_mean, v[5]))
        actual_row += 1

    # FIN SIMULACION

    # ESTILO

    al = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
    fil_header = PatternFill("solid", fgColor="b0e0e6")
    fil_att = PatternFill("solid", fgColor="CCCCCC")
    ws['A1'].fill = fil_header
    ws['A5'].fill = fil_att
    ws['A4'].fill = fil_att
    ws2['A1'].fill = fil_header
    ws2['A4'].fill = fil_att
    ws2['A5'].fill = fil_att

    for l in range(ord('a'), ord('p')):
        ws[chr(l).upper()+str(7)].fill = fil_header
        ws2[chr(l).upper()+str(7)].fill = fil_header

    for l in range(ord('a'), ord('p')):
        ws.column_dimensions[chr(l).upper()].width = 25
        ws2.column_dimensions[chr(l).upper()].width = 25

    ws['A1'] = "Metricas por item (SKU)"
    ws.cell(row=4, column=1, value="Alpha")
    ws.cell(row=4, column=2, value=0.1)
    ws.cell(row=5, column=1, value="Total periodo")
    ws.cell(row=5, column=2, value=total_period)
    ws.cell(row=7, column=1, value="SKU")
    ws.cell(row=7, column=2, value="Demanda Total")
    ws.cell(row=7, column=3, value="Demanda Promedio")
    ws.cell(row=7, column=4, value="Desv. Estándar Demanda")
    ws.cell(row=7, column=5, value="Intervalo Promedio")
    ws.cell(row=7, column=6, value="Primer Pedido")
    ws.cell(row=7, column=7, value="Ultimo Pedido")
    ws.cell(row=7, column=8, value="Tamaño Periodo (dias)")
    ws.cell(row=7, column=9, value="[p(mu)] P Gorro")
    ws.cell(row=7, column=10, value="[z(t)] Z Gorro")
    ws.cell(row=7, column=11, value="[y(t)] Suma demanda")
    ws.cell(row=7, column=12, value="[Lambda] Promedio demanda")
    ws.cell(row=7, column=13, value="Stock Max.")
    ws.cell(row=7, column=14, value="Stock Shortages")
    ws.cell(row=7, column=15, value="Stock medio")
    ws.cell(row=7, column=16, value="MSE(t) medio")
    ws.cell(row=7, column=17, value="MSE(mu) medio")
    ws.cell(row=7, column=17, value="M(mu) medio")

    ws2['A1'] = "Simulacion de Croston"
    ws2.cell(row=7, column=1, value="SKU")
    ws2.cell(row=7, column=2, value="Mu")
    ws2.cell(row=7, column=3, value="Sigma")
    ws2.cell(row=7, column=4, value="P")
    ws2.cell(row=7, column=5, value="R promedio")
    ws2.cell(row=7, column=6, value="NS promedio")
    ws2.cell(row=4, column=1, value="# periodos")
    ws2.cell(row=4, column=2, value=NUMERO_RUNS)
    ws2.cell(row=5, column=1, value="# simulaciones")
    ws2.cell(row=5, column=2, value=NUMERO_SIMU)

    for element in plot_data:
        ws3.append(element)

    chart = BubbleChart()
    chart.title = "p vs NS"
    chart.style = 18
    chart.x_axis.title = 'p'
    chart.y_axis.title = 'NS'

    xvalues = Reference(ws3, min_col=3, min_row=1, max_row=len(plot_data)-1)
    yvalues = Reference(ws3, min_col=2, min_row=1, max_row=len(plot_data)-1)
    size = Reference(ws3, min_col=1, min_row=1, max_row=len(plot_data)-1)
    series = Series(values=yvalues, xvalues=xvalues, zvalues=size, title_from_data=True)
    chart.series.append(series)

    # place the chart starting in cell E1
    ws3.add_chart(chart, "E1")
    wb.save("producto_punto/static/analisis.xlsx")
    xlsx_data = []

    return xlsx_data


def validar_input(contenido):
    ''' Recibe el contenido del archivo y verifica que el
    contenido corresponda al formato requerido.
    Si el formato corresponde, retorna un par (True, elemento)
    donde elemento es un objeto con todos los datos necesarios
    para procesar posteriormente.
    Si el formato no corresponde, retorna un par (False, error)
    donde error es un string indicando cual es el error.'''

    #  Abrimos el archivo como un documento excel
    #  Se utiliza try y except para evitar que si el usuario sube un
    #  archivo corrupto, nuestra aplicación se caiga
    #  Se intenta (try) abrir el archivo, si no se puede, se hace la
    #  excepcion (except) de arrojar un error.

    try:
        book = open_workbook(file_contents=contenido)
    except:
        return (False, 'Archivo corrupto o ilegible, ' +
                'asegúrese de solo modificar los datos de la plantilla.')

    sheet1 = book.sheet_by_index(0)
    sheet2 = book.sheet_by_index(1)

    leadtimes = dict()
    f2 = sheet2.nrows
    c2 = sheet2.ncols
    for f in range(8, f2):
        leadtimes[sheet2.cell(f, 0).value] = sheet2.cell(f, c2-1).value

    dictionary = dict()
    filas = sheet1.nrows
    columnas = sheet1.ncols

    for fila in range(8, filas):
        vector = []
        # Solo usar data cuya frecuencia sea C, D, Z
        if sheet1.cell(fila, 11).value in ['C', 'D', 'Z']:
            for columna in [0, 2, 5, 10, 11, 13]:
                valor_celda = sheet1.cell(fila, columna).value
                vector.append(valor_celda)
            if vector[0] in dictionary:
                dictionary[vector[0]].append(vector)
            else:
                dictionary[vector[0]] = [vector]

    data = WriteToExcel(dictionary, leadtimes)

    return (True, data)
