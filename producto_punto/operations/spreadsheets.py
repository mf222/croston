import random
import numpy as np
from scipy.stats import norm

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.chart import Series, Reference, BubbleChart
from xlrd import open_workbook

from collections import defaultdict
from datetime import datetime, timedelta


def read_sheets(*files):
    for f in files:
        yield open_workbook(f).sheet_by_index(0)


def parse_files(file1, file2):
    sheet1, sheet2 = read_sheets(file1, file2)
    if "demanda" not in sheet1.cell(7, 5).value.lower():
        sheet2, sheet1 = read_sheets(file1, file2)

    leadtimes = defaultdict(lambda: [])
    for f in range(8, sheet2.nrows):
        key = sheet2.cell(f, 0).value
        leadtime = sheet2.cell(f, sheet2.ncols - 1).value
        actual_r = sheet2.cell(f, 10).value
        vector = []
        vector.append(leadtime)
        vector.append(actual_r)
        leadtimes[key].append(vector)

    data = defaultdict(lambda: [])

    for fila in range(8, sheet1.nrows):
        # Solo usar data cuya frecuencia sea C, D, Z
        if sheet1.cell(fila, 11).value in ['C', 'D', 'Z']:
            vector = []
            for columna in [0, 2, 5, 10, 11, 13]:
                valor_celda = sheet1.cell(fila, columna).value
                vector.append(valor_celda)
            data[vector[0]].append(vector)

    return data, leadtimes


def run_operation(data, leadtimes):
    workbook = openpyxl.Workbook()

    # Sheet con datos
    ws1 = workbook.get_active_sheet()

    # Sheet con resultado simulacion
    ws2 = workbook.create_sheet(title="Croston")

    # Sheet para el gráfico
    # ws3 = workbook.create_sheet(title="p vs NS")

    # Definir aqui el numero de veces que correra
    NUMERO_RUNS = 100
    NUMERO_SIMU = 100

    # Para determinar cuanto es el intervalo de días que se simula
    min_date = None
    max_date = None

    # Diccionario para los datos de la simulación
    data_simulation = dict()

    # Lista para el grafico
    plot_data = []

    actual_row = 7
    for k, v in data.items():
        # Solo se simulan datos que tengan más de 1 pedido
        pedidos = len(v)
        if pedidos > 1:
            data_elements = []
            alldda = []
            alldates = []
            no_null = 0

            for element in v:
                if element[2] > 0:
                    dda = element[2]
                    no_null += 1
                    alldda.append(dda)
                    alldates.append(datetime.strptime(element[1], '%d-%m-%Y'))

            # Para determinar los periodos e intervalos
            alldates.sort()
            difference_days = []
            for i in range(0, len(alldates)-1):
                difference_days.append((alldates[i+1]-alldates[i]).days)
            period = (alldates[-1]-alldates[0]).days
            m_interval = np.mean(difference_days)

            # Para determinar demanda media y desv. estandar
            m_dda = np.mean(alldda)
            std_dda = np.std(alldda)

            if m_interval == 0:
                m_interval = (max_date-min_date).days

            # CROSTON METRICAS
            p_croston = m_interval/no_null
            q_croston = 1
            z_croston = m_dda
            sigma_croston = std_dda

            y_t = []
            y_t.append(z_croston/p_croston)
            y_t_d = []
            y_t_d_pow = []

            e_mu = []
            e_mu_pow = []

            Mn = []
            Mn.append(std_dda)

            Rt = []
            Rt.append(m_dda+m_dda*std_dda)

            shortage = []
            Mn_temp = []

            alpha = 0.1
            actual_dda = 0
            actual_day = alldates[0]

            r_actual = 0
            low_ns = 0
            if k in leadtimes:

                r_actual = data_elements.append(leadtimes.get(k)[0])

            if r_actual is None:
                r_actual = 0
            for i in range(period):
                e_val = 0
                y_t_d_val = 0 - z_croston/p_croston
                mn_val = Mn[i]
                shortage_val = 0
                mn_temp = 0
                if actual_day == datetime.strptime(v[actual_dda][1], '%d-%m-%Y'):
                    e_val = v[actual_dda][2] - z_croston
                    y_t_d_val = v[actual_dda][2] - z_croston/p_croston
                    if y_t_d_val > r_actual:
                        low_ns += 1
                    p_croston = (p_croston*(1-alpha))+(alpha*q_croston)
                    z_croston = (z_croston*(1-alpha))+(alpha*v[actual_dda][2])
                    q_croston = 0
                    mn_val = Mn[i]*(1-Mn[i]) + Mn[i]*abs(y_t_d_val)
                    actual_dda += 1
                    mn_temp = Mn[i]
                    sigma_croston = sigma_croston*(1-alpha)+alpha*abs(e_val)
                if Rt[i] < v[actual_dda][2]:
                    shortage_val = 1
                q_croston += 1
                actual_day = actual_day + timedelta(days=1)
                y_t.append(z_croston/p_croston)
                e_mu.append(e_val)
                e_mu_pow.append(e_val**2)
                y_t_d.append(y_t_d_val)
                y_t_d_pow.append(y_t_d_val**2)
                Mn.append(mn_val)
                Rt.append(np.floor(z_croston+z_croston*mn_val))
                shortage.append(shortage_val)
                Mn_temp.append(mn_temp)

            e_nzero = np.trim_zeros(e_mu_pow)
            mn_nzero = np.trim_zeros(Mn_temp)

            if len(e_nzero) == 0:
                e_nzero.append(0)
            if len(mn_nzero) == 0:
                mn_nzero.append(0)


            data_elements.append(pedidos)  # 0 : TOTAL PEDIDOS
            data_elements.append(period/len(alldda))  # 1 : INTERVALO, P AVG
            data_elements.append(m_dda)  # 2 : DEMANDA MEDIA, MU AVG
            # data_elements.append(std_dda)  # 3 : SIGMA
            data_elements.append(y_t[-1])  # 4 : MU MEAN - MU
            data_elements.append(Mn[-1])  # 5 : SIGMA ?
            data_elements.append(p_croston)  # 6 : P
            data_elements.append(std_dda)  # 7
            data_elements.append(z_croston)  # 8 : Z croston
            data_elements.append(sigma_croston)  # 9

            if k in leadtimes:
                data_elements.append(leadtimes.get(k)[0][0])  # 10 : leadtimes
                data_elements.append(leadtimes.get(k)[0][1])  # 11 : R actual
                data_elements.append(low_ns)  # 12 : NS actual
            else:
                data_elements.append(None)
                data_elements.append(None)
                data_elements.append(None)

            data_simulation[k] = data_elements
            actual_row += 1

            if min_date is None:
                min_date = alldates[0]
            elif alldates[0] < min_date:
                min_date = alldates[0]

            if max_date is None:
                max_date = alldates[-1]
            elif alldates[-1] > max_date:
                max_date = alldates[-1]

            ws1.cell(row=actual_row, column=1, value=k)
            ws1.cell(row=actual_row, column=2, value=np.sum(alldda))
            ws1.cell(row=actual_row, column=3, value=round(m_dda, 0))
            ws1.cell(row=actual_row, column=4, value=round(std_dda, 2))
            ws1.cell(row=actual_row, column=5, value=round(m_interval, 0))
            ws1.cell(row=actual_row, column=6, value=("{}/{}/{}".format(alldates[0].timetuple()[2], alldates[0].timetuple()[1], alldates[0].timetuple()[0])))
            ws1.cell(row=actual_row, column=7, value=("{}/{}/{}".format(alldates[-1].timetuple()[2], alldates[-1].timetuple()[1], alldates[-1].timetuple()[0])))
            ws1.cell(row=actual_row, column=8, value=period)
            ws1.cell(row=actual_row, column=9, value=round(p_croston, 2))
            ws1.cell(row=actual_row, column=10, value=round(z_croston, 2))
            ws1.cell(row=actual_row, column=11, value=round(np.sum(y_t), 2))
            ws1.cell(row=actual_row, column=12, value=round(np.mean(y_t), 2))
            ws1.cell(row=actual_row, column=13, value=max(Rt))
            ws1.cell(row=actual_row, column=14, value=np.sum(shortage))
            ws1.cell(row=actual_row, column=15, value=np.mean(Rt))
            ws1.cell(row=actual_row, column=16, value=np.mean(y_t_d_pow))
            ws1.cell(row=actual_row, column=17, value=np.mean(e_nzero))
            ws1.cell(row=actual_row, column=18, value=np.mean(mn_nzero))


    total_period = (max_date - min_date).days

    # INICIO SIMULACION
    actual_row = 8
    for k, v in data_simulation.items():

        ALL_NS = []
        ALL_R_0 = []
        leadtime = 3
        actual_r = "R actual no encontrado en segundo archivo"
        actual_ns = "R actual no encontrano en segundo archivo"

        if v[10] is not None:
            leadtime = v[10]


            if v[11] is not '':
                actual_ns = 1 - v[10]/v[0]
                actual_r = v[11]
            if actual_ns.isnumeric():
                if int(actual_ns) <= 0:
                    actual_ns = 0
                elif int(actual_ns) >= 0:
                    actual_ns = str(round(actual_ns, 2)*100) + "%"

        for simulations in range(10):
            NS = 0
            R_0 = 0
            while NS < 0.8:
                dda = []
                dda_desp = []
                R = []
                R_0 += 1
                R.append(R_0)
                y_more_R = []
                for i in range(NUMERO_RUNS):
                    selector = 1/(v[6])
                    y = 0
                    if selector <= random.random():
                        y = np.floor(norm.ppf(random.random(), v[8], v[9]))
                        if y < 1:
                            y = 1
                    dda.append(y)
                for i in range(NUMERO_RUNS):
                    desplazamiento = i - leadtime
                    if desplazamiento < 0:
                        dda_desp.append(0)
                    else:
                        dda_desp.append(dda[i])

                for i in range(1, NUMERO_RUNS):
                    R_i = (R[i-1]-dda[i])+dda_desp[i]
                    R.append(R_i)

                for i in range(NUMERO_RUNS):
                    ymr = 0
                    if dda[i] > R[i]:
                        ymr = 1
                    y_more_R.append(ymr)

                numero_dda = 0
                for i in range(len(dda)):
                    if dda[i] != 0:
                        numero_dda += 1
                NS = 1 - (np.sum(y_more_R)/numero_dda)

            ALL_NS.append(NS)
            ALL_R_0.append(R_0)
        ns_mean = np.mean(NS)

        ws2.cell(row=actual_row, column=1, value=k)
        ws2.cell(row=actual_row, column=2, value=round(v[3], 2))
        ws2.cell(row=actual_row, column=3, value=round(v[7], 2))
        ws2.cell(row=actual_row, column=4, value=round(v[6], 2))
        ws2.cell(row=actual_row, column=5, value=round(v[8], 2))
        ws2.cell(row=actual_row, column=6, value=round(v[9], 2))
        ws2.cell(row=actual_row, column=7, value=round(v[6], 2))
        ws2.cell(row=actual_row, column=8, value=actual_r)
        ws2.cell(row=actual_row, column=9, value=round(np.mean(ALL_R_0), 0))
        ws2.cell(row=actual_row, column=10, value=actual_ns)
        ws2.cell(row=actual_row, column=11, value="{}%".format(round(ns_mean, 2)*100))
        plot_data.append((1, ns_mean, v[5]))
        actual_row += 1

    # FIN SIMULACION

    # ESTILO
    fil_header = PatternFill("solid", fgColor="b0e0e6")
    fil_att = PatternFill("solid", fgColor="CCCCCC")
    ws1['A1'].fill = fil_header
    ws1['A5'].fill = fil_att
    ws1['A4'].fill = fil_att
    ws2['A1'].fill = fil_header
    ws2['A4'].fill = fil_att
    ws2['A5'].fill = fil_att

    for l in range(ord('a'), ord('p')):
        ws1[chr(l).upper()+str(7)].fill = fil_header
        ws2[chr(l).upper()+str(7)].fill = fil_header

    for l in range(ord('a'), ord('p')):
        ws1.column_dimensions[chr(l).upper()].width = 15
        ws2.column_dimensions[chr(l).upper()].width = 15

    ws1.column_dimensions['A'].width = 20
    ws2.column_dimensions['A'].width = 20

    ws1['A1'] = "Metricas por item (SKU)"
    ws1.cell(row=4, column=1, value="Alpha")
    ws1.cell(row=4, column=2, value=0.1)
    ws1.cell(row=5, column=1, value="Total periodo")
    ws1.cell(row=5, column=2, value=total_period)
    ws1.cell(row=7, column=1, value="SKU")
    ws1.cell(row=7, column=2, value="Demanda Total")
    ws1.cell(row=7, column=3, value="Demanda Promedio")
    ws1.cell(row=7, column=4, value="Desv. Estándar Demanda")
    ws1.cell(row=7, column=5, value="Intervalo entre Pedidos Promedio (dias)")
    ws1.cell(row=7, column=6, value="Primer Pedido")
    ws1.cell(row=7, column=7, value="Ultimo Pedido")
    ws1.cell(row=7, column=8, value="Intervalo entre Primer y Ultimo Pedido (dias)")
    ws1.cell(row=7, column=9, value="[p(mu)] P Gorro")
    ws1.cell(row=7, column=10, value="[z(t)] Z Gorro")
    ws1.cell(row=7, column=11, value="[y(t)] Suma demanda")
    ws1.cell(row=7, column=12, value="[Lambda] Promedio demanda")
    ws1.cell(row=7, column=13, value="Stock Max.")
    ws1.cell(row=7, column=14, value="Stock Shortages")
    ws1.cell(row=7, column=15, value="Stock medio")
    ws1.cell(row=7, column=16, value="MSE(t) medio")
    ws1.cell(row=7, column=17, value="MSE(mu) medio")
    ws1.cell(row=7, column=17, value="M(mu) medio")

    ws2['A1'] = "Simulacion de Croston"
    ws2.cell(row=7, column=1, value="SKU")
    ws2.cell(row=7, column=2, value="Mu AVG [Muestra]")
    ws2.cell(row=7, column=3, value="Sigma [Muestra]")
    ws2.cell(row=7, column=4, value="P AVG [Muestra]")
    ws2.cell(row=7, column=5, value="Mu [Croston]")
    ws2.cell(row=7, column=6, value="Sigma [Croston]")
    ws2.cell(row=7, column=7, value="P [Croston]")
    ws2.cell(row=7, column=8, value="R Actual")
    ws2.cell(row=7, column=9, value="R Recomendado")
    ws2.cell(row=7, column=10, value="NS Actual")
    ws2.cell(row=7, column=11, value="NS Recomendado")
    ws2.cell(row=4, column=1, value="# periodos")
    ws2.cell(row=4, column=2, value=NUMERO_RUNS)
    ws2.cell(row=5, column=1, value="# simulaciones")
    ws2.cell(row=5, column=2, value=NUMERO_SIMU)

    # for element in plot_data:
    #     ws3.append(element)

    # chart = BubbleChart()
    # chart.title = "p vs NS"
    # chart.style = 18
    # chart.x_axis.title = 'p'
    # chart.y_axis.title = 'NS'
    #
    # xvalues = Reference(ws3, min_col=3, min_row=1, max_row=len(plot_data)-1)
    # yvalues = Reference(ws3, min_col=2, min_row=1, max_row=len(plot_data)-1)
    # size = Reference(ws3, min_col=1, min_row=1, max_row=len(plot_data)-1)
    # series = Series(values=yvalues, xvalues=xvalues, zvalues=size,
    #                 title_from_data=True)
    # chart.series.append(series)

    # place the chart starting in cell E1
    # ws3.add_chart(chart, "E1")
    return workbook
